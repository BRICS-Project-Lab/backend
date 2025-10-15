
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings
from django.utils.translation import gettext as _
from django.http import HttpResponse
import logging
import csv
import json
from datetime import datetime

logger = logging.getLogger(__name__)

def send_notification_email(user, subject, template_name, context=None):
    """
    Отправка уведомлений по email
    
    Args:
        user: объект пользователя
        subject: тема письма
        template_name: имя шаблона без расширения
        context: дополнительный контекст для шаблона
    
    Returns:
        bool: True если письмо отправлено, False если ошибка
    """
    if not user.email:
        logger.warning(f"User {user.username} has no email address")
        return False
    
    try:
        context = context or {}
        context['user'] = user
        context['site_name'] = 'BRICS AI Registry'
        context['site_url'] = settings.SITE_URL if hasattr(settings, 'SITE_URL') else 'http://localhost:8000'
        
        # Генерируем HTML и текстовую версию
        html_message = render_to_string(f'emails/{template_name}.html', context)
        text_message = render_to_string(f'emails/{template_name}.txt', context)
        
        send_mail(
            subject=subject,
            message=text_message,
            html_message=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=False,
        )
        
        logger.info(f"Email sent to {user.email}: {subject}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send email to {user.email}: {str(e)}")
        return False

def log_user_action(user, action, obj=None, comment=None, request=None):
    """
    Логирование действий пользователя
    
    Args:
        user: пользователь, выполняющий действие
        action: тип действия (из AuditLog.Action)
        obj: объект, над которым выполняется действие
        comment: комментарий
        request: HTTP request для получения IP и user-agent
    """
    from apps.common.models import AuditLog
    from django.contrib.contenttypes.models import ContentType
    
    try:
        audit_data = {
            'action': action,
            'performed_by': user,
            'comment': comment or '',
        }
        
        if obj:
            audit_data.update({
                'content_type': ContentType.objects.get_for_model(obj),
                'object_id': obj.pk,
            })
        
        if request:
            audit_data.update({
                'ip_address': getattr(request, 'ip_address', None),
                'user_agent': getattr(request, 'user_agent', ''),
            })
        
        AuditLog.objects.create(**audit_data)
        logger.info(f"Action logged: {action} by {user.username}")
        
    except Exception as e:
        logger.error(f"Failed to log user action: {str(e)}")

def get_client_ip(request):
    """
    Получение IP адреса клиента из request
    
    Args:
        request: HTTP request
    
    Returns:
        str: IP адрес клиента
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def export_to_csv(queryset, fields, filename='export.csv'):
    """
    Экспорт QuerySet в CSV файл
    
    Args:
        queryset: Django QuerySet для экспорта
        fields: список кортежей (field_name, header_name) или просто имен полей
        filename: имя файла для скачивания
    
    Returns:
        HttpResponse: CSV файл для скачивания
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Добавляем BOM для корректного отображения в Excel
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Заголовки
    headers = []
    field_names = []
    
    for field in fields:
        if isinstance(field, tuple):
            field_names.append(field[0])
            headers.append(field[1])
        else:
            field_names.append(field)
            headers.append(field.replace('_', ' ').title())
    
    writer.writerow(headers)
    
    # Данные
    for obj in queryset:
        row = []
        for field_name in field_names:
            # Поддержка вложенных полей через точку
            if '.' in field_name:
                value = obj
                for attr in field_name.split('.'):
                    value = getattr(value, attr, None)
                    if value is None:
                        break
            else:
                value = getattr(obj, field_name, None)
            
            # Обработка callable
            if callable(value):
                value = value()
            
            # Конвертация в строку
            if value is None:
                row.append('')
            elif isinstance(value, (list, dict)):
                row.append(json.dumps(value, ensure_ascii=False))
            elif isinstance(value, datetime):
                row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                row.append(str(value))
        
        writer.writerow(row)
    
    return response

def export_to_xlsx(queryset, fields, filename='export.xlsx'):
    """
    Экспорт QuerySet в XLSX файл
    
    Args:
        queryset: Django QuerySet для экспорта
        fields: список кортежей (field_name, header_name) или просто имен полей
        filename: имя файла для скачивания
    
    Returns:
        HttpResponse: XLSX файл для скачивания
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import io
    except ImportError:
        logger.error("openpyxl not installed, cannot export to XLSX")
        # Fallback to CSV
        return export_to_csv(queryset, fields, filename.replace('.xlsx', '.csv'))
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Export"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Заголовки
    headers = []
    field_names = []
    
    for field in fields:
        if isinstance(field, tuple):
            field_names.append(field[0])
            headers.append(field[1])
        else:
            field_names.append(field)
            headers.append(field.replace('_', ' ').title())
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Данные
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(field_names, 1):
            # Поддержка вложенных полей
            if '.' in field_name:
                value = obj
                for attr in field_name.split('.'):
                    value = getattr(value, attr, None)
                    if value is None:
                        break
            else:
                value = getattr(obj, field_name, None)
            
            # Обработка callable
            if callable(value):
                value = value()
            
            # Запись значения
            if value is None:
                cell_value = ''
            elif isinstance(value, (list, dict)):
                cell_value = json.dumps(value, ensure_ascii=False)
            elif isinstance(value, datetime):
                cell_value = value
            else:
                cell_value = value
            
            worksheet.cell(row=row_num, column=col_num, value=cell_value)
    
    # Автоширина колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Сохранение в память
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def export_queryset_to_csv(queryset, fields, filename='export.csv'):
    """Alias для обратной совместимости"""
    return export_to_csv(queryset, fields, filename)

def export_queryset_to_xlsx(queryset, fields, filename='export.xlsx'):
    """Alias для обратной совместимости"""
    return export_to_xlsx(queryset, fields, filename)

def format_file_size(size_bytes):
    """
    Форматирование размера файла в человекочитаемый вид
    
    Args:
        size_bytes: размер в байтах
    
    Returns:
        str: отформатированный размер (например, "1.5 MB")
    """
    if not size_bytes:
        return "0 B"
    
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    unit_index = 0
    size = float(size_bytes)
    
    while size >= 1024.0 and unit_index < len(units) - 1:
        size /= 1024.0
        unit_index += 1
    
    return f"{size:.1f} {units[unit_index]}"

def truncate_text(text, max_length=100, suffix='...'):
    """
    Обрезка текста с добавлением суффикса
    
    Args:
        text: исходный текст
        max_length: максимальная длина
        suffix: суффикс для добавления в конец
    
    Returns:
        str: обрезанный текст
    """
    if not text or len(text) <= max_length:
        return text
    
    return text[:max_length - len(suffix)] + suffix

def generate_slug(text, max_length=50):
    """
    Генерация slug из текста
    
    Args:
        text: исходный текст
        max_length: максимальная длина slug
    
    Returns:
        str: slug
    """
    from django.utils.text import slugify
    import re
    
    # Транслитерация для кириллицы
    translit_dict = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya'
    }
    
    text_lower = text.lower()
    transliterated = ''
    
    for char in text_lower:
        transliterated += translit_dict.get(char, char)
    
    slug = slugify(transliterated)
    return slug[:max_length]

def validate_json_field(value, expected_type=None):
    """
    Валидация JSON поля
    
    Args:
        value: значение для валидации
        expected_type: ожидаемый тип (dict, list, и т.д.)
    
    Returns:
        bool: True если валидно, False иначе
    """
    if value is None:
        return True
    
    if expected_type:
        return isinstance(value, expected_type)
    
    return isinstance(value, (dict, list))

def get_or_none(model, **kwargs):
    """
    Получить объект или None вместо исключения
    
    Args:
        model: Django модель
        **kwargs: параметры фильтрации
    
    Returns:
        object или None
    """
    try:
        return model.objects.get(**kwargs)
    except model.DoesNotExist:
        return None
    except model.MultipleObjectsReturned:
        logger.warning(f"Multiple {model.__name__} objects returned for {kwargs}")
        return model.objects.filter(**kwargs).first()

def batch_create(model, data_list, batch_size=1000):
    """
    Массовое создание объектов батчами
    
    Args:
        model: Django модель
        data_list: список словарей с данными
        batch_size: размер батча
    
    Returns:
        int: количество созданных объектов
    """
    objects = [model(**data) for data in data_list]
    created = 0
    
    for i in range(0, len(objects), batch_size):
        batch = objects[i:i + batch_size]
        model.objects.bulk_create(batch)
        created += len(batch)
    
    logger.info(f"Created {created} {model.__name__} objects")
    return created

def send_admin_notification(subject, message, level='info'):
    """
    Отправка уведомления всем администраторам
    
    Args:
        subject: тема уведомления
        message: текст уведомления
        level: уровень важности (info, warning, error)
    """
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    admins = User.objects.filter(role='admin', is_active=True, is_blocked=False)
    
    for admin in admins:
        try:
            send_notification_email(
                user=admin,
                subject=subject,
                template_name='admin_notification',
                context={
                    'message': message,
                    'level': level
                }
            )
        except Exception as e:
            logger.error(f"Failed to send notification to admin {admin.username}: {str(e)}")

def clean_html(html_text):
    """
    Очистка HTML от потенциально опасных тегов
    
    Args:
        html_text: HTML текст
    
    Returns:
        str: очищенный текст
    """
    import re
    
    if not html_text:
        return ''
    
    # Удаляем script и style теги
    html_text = re.sub(r'<script[^>]*>.*?</script>', '', html_text, flags=re.DOTALL)
    html_text = re.sub(r'<style[^>]*>.*?</style>', '', html_text, flags=re.DOTALL)
    
    # Удаляем потенциально опасные атрибуты
    html_text = re.sub(r'on\w+="[^"]*"', '', html_text)
    html_text = re.sub(r"on\w+='[^']*'", '', html_text)
    
    return html_text