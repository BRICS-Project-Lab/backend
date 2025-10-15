from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from django.http import HttpResponse
from django.core.serializers import serialize
import json
import csv
import io
from apps.ai_modules.models import AIModule
from apps.tags.models import Tag
from apps.publications.models import Publication
from apps.common.utils import export_to_csv, export_to_xlsx

class ModulesExportView(APIView):
    """Экспорт модулей"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        format_type = request.query_params.get('format', 'json').lower()
        
        # Фильтрация
        queryset = AIModule.objects.filter(status=AIModule.Status.ACTIVE)
        
        country = request.query_params.get('country')
        if country:
            queryset = queryset.filter(country=country)
        
        tags = request.query_params.getlist('tags')
        if tags:
            queryset = queryset.filter(aimoduletag__tag__id__in=tags).distinct()
        
        if format_type == 'json':
            return self._export_json(queryset)
        elif format_type == 'csv':
            return self._export_csv(queryset)
        elif format_type == 'xlsx':
            return self._export_xlsx(queryset)
        else:
            return Response(
                {'error': 'Unsupported format'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def _export_json(self, queryset):
        """Экспорт в JSON"""
        data = []
        for module in queryset:
            module_data = {
                'id': str(module.id),
                'name': module.name,
                'company': module.company,
                'country': module.country,
                'params_count': module.params_count,
                'description': module.task_short_description,
                'version': module.version,
                'license': module.license_type,
                'created_at': module.created_at.isoformat(),
                'tags': [tag.name for tag in module.get_tags()],
                'like_count': module.get_like_count(),
                'publications_count': module.publications.count()
            }
            data.append(module_data)
        
        response = HttpResponse(
            json.dumps(data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="ai_modules.json"'
        return response
    
    def _export_csv(self, queryset):
        """Экспорт в CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="ai_modules.csv"'
        
        writer = csv.writer(response)
        
        # Заголовки
        writer.writerow([
            'ID', 'Name', 'Company', 'Country', 'Parameters Count',
            'Description', 'Version', 'License', 'Created At', 'Tags', 'Likes'
        ])
        
        # Данные
        for module in queryset:
            writer.writerow([
                str(module.id),
                module.name,
                module.company,
                module.country,
                module.params_count,
                module.task_short_description,
                module.version,
                module.license_type,
                module.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                ', '.join([tag.name for tag in module.get_tags()]),
                module.get_like_count()
            ])
        
        return response
    
    def _export_xlsx(self, queryset):
        """Экспорт в XLSX"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return Response(
                {'error': 'XLSX export not available'}, 
                status=status.HTTP_501_NOT_IMPLEMENTED
            )
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "AI Modules"
        
        # Заголовки
        headers = [
            'ID', 'Name', 'Company', 'Country', 'Parameters Count',
            'Description', 'Version', 'License', 'Created At', 'Tags', 'Likes'
        ]
        
        # Стилизация заголовков
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Данные
        for row_num, module in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1, value=str(module.id))
            worksheet.cell(row=row_num, column=2, value=module.name)
            worksheet.cell(row=row_num, column=3, value=module.company)
            worksheet.cell(row=row_num, column=4, value=module.country)
            worksheet.cell(row=row_num, column=5, value=module.params_count)
            worksheet.cell(row=row_num, column=6, value=module.task_short_description)
            worksheet.cell(row=row_num, column=7, value=module.version)
            worksheet.cell(row=row_num, column=8, value=module.license_type)
            worksheet.cell(row=row_num, column=9, value=module.created_at)
            worksheet.cell(row=row_num, column=10, value=', '.join([tag.name for tag in module.get_tags()]))
            worksheet.cell(row=row_num, column=11, value=module.get_like_count())
        
        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column.column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Сохранение в память
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="ai_modules.xlsx"'
        
        return response

class TagsExportView(APIView):
    """Экспорт тегов"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        tags = Tag.objects.filter(is_active=True).select_related('category').annotate(
            usage_count=Count('aimoduletag')
        )
        
        data = []
        for tag in tags:
            data.append({
                'id': tag.id,
                'name': tag.name,
                'category': tag.category.name,
                'description': tag.description,
                'color': tag.get_color_or_default(),
                'usage_count': tag.usage_count,
                'created_at': tag.created_at.isoformat()
            })
        
        response = HttpResponse(
            json.dumps(data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="tags.json"'
        return response

class PublicationsExportView(APIView):
    """Экспорт публикаций"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        publications = Publication.objects.select_related('ai_module')
        
        data = []
        for pub in publications:
            data.append({
                'id': pub.id,
                'title': pub.title,
                'authors': pub.authors,
                'journal': pub.journal_conference,
                'publication_date': pub.publication_date.isoformat(),
                'doi': pub.doi,
                'url': pub.url,
                'abstract': pub.abstract,
                'keywords': pub.keywords,
                'citation_count': pub.citation_count,
                'ai_module': pub.ai_module.name,
                'ai_module_company': pub.ai_module.company
            })
        
        response = HttpResponse(
            json.dumps(data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="publications.json"'
        return response

class StatsExportView(APIView):
    """Экспорт общей статистики"""
    permission_classes = [permissions.IsAdminUser]
    
    def get(self, request):
        from .analytics_views import OverviewAnalyticsView
        
        # Используем существующую аналитику
        analytics_view = OverviewAnalyticsView()
        analytics_view.request = request
        analytics_response = analytics_view.get(request)
        
        response = HttpResponse(
            json.dumps(analytics_response.data, indent=2, ensure_ascii=False),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="system_stats.json"'
        return response


class ModulesCSVExport(APIView):
    """Экспорт модулей в CSV"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        """Экспорт в CSV"""
        # Фильтрация
        queryset = AIModule.objects.filter(status=AIModule.Status.ACTIVE).select_related('created_by')
        
        country = request.query_params.get('country')
        if country:
            queryset = queryset.filter(country=country)
        
        tags = request.query_params.getlist('tags')
        if tags:
            queryset = queryset.filter(aimoduletag__tag__id__in=tags).distinct()
        
        # Определяем поля для экспорта
        fields = [
            ('id', 'ID'),
            ('name', 'Name'),
            ('company', 'Company'),
            ('country', 'Country'),
            ('params_count', 'Parameters Count'),
            ('task_short_description', 'Description'),
            ('version', 'Version'),
            ('license_type', 'License'),
            ('created_by.username', 'Created By'),
            ('created_at', 'Created At'),
            ('status', 'Status'),
        ]
        
        return export_to_csv(queryset, fields, 'ai_modules.csv')

class ModulesXLSXExport(APIView):
    """Экспорт модулей в XLSX"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        """Экспорт в XLSX"""
        # Фильтрация
        queryset = AIModule.objects.filter(status=AIModule.Status.ACTIVE).select_related('created_by')
        
        country = request.query_params.get('country')
        if country:
            queryset = queryset.filter(country=country)
        
        tags = request.query_params.getlist('tags')
        if tags:
            queryset = queryset.filter(aimoduletag__tag__id__in=tags).distinct()
        
        # Определяем поля для экспорта
        fields = [
            ('id', 'ID'),
            ('name', 'Name'),
            ('company', 'Company'),
            ('country', 'Country'),
            ('params_count', 'Parameters Count'),
            ('task_short_description', 'Description'),
            ('version', 'Version'),
            ('license_type', 'License'),
            ('created_by.username', 'Created By'),
            ('created_at', 'Created At'),
            ('status', 'Status'),
            ('get_like_count', 'Likes'),
        ]
        
        return export_to_xlsx(queryset, fields, 'ai_modules.xlsx')